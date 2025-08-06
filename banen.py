import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta


# =============================================================================
# CONFIGURATION
# =============================================================================
input_file = 'Geboekte producten.xlsx'          # Your reservation Excel file
output_file_full = 'BanenPlanning.xlsx'   # Full schedule


# =============================================================================
# STEP 1: LOAD DATA
# =============================================================================
df = pd.read_excel(input_file)


# =============================================================================
# STEP 2: PROCESS DATA
# =============================================================================

def calculate_lanes_needed(row):
    """
    Calculate the number of lanes needed based on the number of people.
    - If 4+ people: 6 persons per lane (rounded up)
    - If < 4 people: number of people equals number of lanes
    """
    if row['Aantal'] >= 4:
        return (row['Aantal'] + 5) // 6  # 6 persons per lane, round up
    else:
        return row['Aantal']  # Aantal is lanes if < 4

df['lanes_needed'] = df.apply(calculate_lanes_needed, axis=1)

# Initialize lane booking memory
lanes = {i: [] for i in range(1, 9)}  # lanes 1 to 8
assignments = []

# Sort by group and start time
df = df.sort_values(by=['Begindatum', 'Groep'])  # Sort by time first, then group


def is_lane_free(lane, new_start, new_end):
    """
    Helper function to check if a lane is free during the specified time period.
    """
    for booked_start, booked_end in lanes[lane]:
        # Check for overlap: new booking overlaps if it starts before existing ends AND ends after existing starts
        if new_start < booked_end and new_end > booked_start:
            return False
    return True


# =============================================================================
# STEP 3: ASSIGN LANES TO RESERVATIONS
# =============================================================================

# Facing pairs configuration
facing_pairs = [(1, 2), (3, 4), (5, 6), (7, 8)]
previous_assignment = {}

# Group bookings by time slot for better processing
df_grouped = df.groupby('Begindatum')

# Process each time slot
for time_slot, time_group in df_grouped:
    start_time = pd.to_datetime(time_slot)
    print(f"\n⏰ Processing time slot: {start_time}")
    
    # Separate consecutive bookings from new bookings
    consecutive_bookings = []
    new_bookings = []
    
    for idx, row in time_group.iterrows():
        groep = row['Groep']
        lanes_needed = row['lanes_needed']
        previous_hour = start_time - timedelta(hours=1)
        
        # Check if this is a consecutive booking
        is_consecutive = False
        for assignment in assignments:
            if (assignment['Groep'] == groep and 
                assignment['Starttijd'] == previous_hour):
                is_consecutive = True
                break
        
        booking_info = {
            'idx': idx,
            'groep': groep,
            'lanes_needed': lanes_needed,
            'start_time': start_time,
            'end_time': start_time + timedelta(minutes=55)
        }
        
        if is_consecutive:
            consecutive_bookings.append(booking_info)
        else:
            new_bookings.append(booking_info)
    
    print(f"  📅 Consecutive bookings: {[b['groep'] for b in consecutive_bookings]}")
    print(f"  🆕 New bookings: {[b['groep'] for b in new_bookings]}")
    
    # Process consecutive bookings FIRST
    all_bookings = consecutive_bookings + new_bookings
    
    for booking in all_bookings:
        groep = booking['groep']
        lanes_needed = booking['lanes_needed']
        start_time = booking['start_time']
        end_time = booking['end_time']
        
        # Determine possible lanes based on start time
        if start_time.minute == 0:
            possible_lanes = range(1, 5)  # lanes 1-4
            possible_pairs = [(1, 2), (3, 4)]
        elif start_time.minute == 30:
            possible_lanes = range(5, 9)  # lanes 5-8
            possible_pairs = [(5, 6), (7, 8)]
        else:
            print(f"Skipping {groep} at {start_time} — invalid start time (must be :00 or :30).")
            continue

        assigned_lanes = []

        # Check if this is a consecutive booking
        if booking in consecutive_bookings:
            # PRIORITY 1: Handle consecutive booking
            previous_hour = start_time - timedelta(hours=1)
            
            # Find the previous assignment
            for assignment in assignments:
                if (assignment['Groep'] == groep and 
                    assignment['Starttijd'] == previous_hour):
                    previous_lanes = assignment['Lanes']
                    
                    if len(previous_lanes) == lanes_needed:
                        # Check if the same lanes are available
                        all_lanes_free = True
                        for lane in previous_lanes:
                            if not is_lane_free(lane, start_time, end_time):
                                all_lanes_free = False
                                break
                        
                        if all_lanes_free:
                            assigned_lanes = previous_lanes
                            print(f"✅ Continuing {groep} on same lanes {assigned_lanes} from previous hour")
                        else:
                            print(f"⚠️ Cannot continue {groep} on same lanes - some lanes not available")
                    else:
                        print(f"⚠️ Cannot continue {groep} - different number of lanes needed")
                    break

        # PRIORITY 2: If no lanes assigned yet, use normal assignment logic
        if not assigned_lanes:
            # First, try full facing pairs
            for pair in possible_pairs:
                if all(is_lane_free(lane, start_time, end_time) for lane in pair):
                    if lanes_needed == 2:
                        assigned_lanes = list(pair)
                        break
                    elif lanes_needed > 2:
                        assigned_lanes.extend(list(pair))
            
            # If still need more lanes (e.g. lanes_needed > 2)
            if len(assigned_lanes) < lanes_needed:
                for lane in possible_lanes:
                    if lane not in assigned_lanes and is_lane_free(lane, start_time, end_time):
                        assigned_lanes.append(lane)
                    if len(assigned_lanes) == lanes_needed:
                        break

        # Check if enough lanes were assigned
        if len(assigned_lanes) < lanes_needed:
            print(f"⚠️ Not enough lanes available for {groep} at {start_time}. Only {len(assigned_lanes)} assigned.")
            continue

        # Book the lanes
        for lane in assigned_lanes:
            lanes[lane].append((start_time, end_time))

        assignments.append({
            'Groep': groep,
            'Starttijd': start_time,
            'Eindtijd': end_time,
            'Lanes': assigned_lanes
        })

        previous_assignment[groep] = (end_time, assigned_lanes)


# =============================================================================
# STEP 4: BUILD THE SCHEDULE DICTIONARY
# =============================================================================

# Build time slots: from 13:00 to 22:00 in half-hour increments
start_time_of_day = datetime.strptime("13:00", "%H:%M")
timeslots = [start_time_of_day + timedelta(minutes=30*i) for i in range(19)]  # 13:00 to 22:00

# Create base schedule
schedule = {}
for time in timeslots:
    schedule[time.strftime("%H:%M")] = {lane: "" for lane in range(1, 9)}

# Fill reservations
for entry in assignments:
    groep = entry['Groep']
    start_time = entry['Starttijd']
    lanes = entry['Lanes']
    time_str = start_time.strftime("%H:%M")
    
    if time_str in schedule:
        for lane in lanes:
            schedule[time_str][lane] = groep
    else:
        print(f"⚠️ Warning: Time {time_str} not in basic schedule range.")


# =============================================================================
# STEP 5: SAVE FULL VERSION (EVEN EMPTY SLOTS)
# =============================================================================

def format_worksheet(ws, schedule_data, title):
    """
    Apply formatting to make the Excel worksheet more readable.
    """
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    time_font = Font(bold=True, size=14)  # Made bigger
    time_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    lane_font = Font(size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Time column - made wider for bigger text
    for i in range(2, 10):  # Lane columns (B-I)
        ws.column_dimensions[chr(64 + i)].width = 18  # Made wider for wrapped text
    
    # Set row heights to accommodate wrapped text
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 25
    
    # Format headers
    for col in range(1, 10):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Format data rows
    for row_idx in range(2, ws.max_row + 1):
        # Format time column with alignment based on time
        time_cell = ws.cell(row=row_idx, column=1)
        time_value = time_cell.value
        time_cell.font = time_font
        time_cell.fill = time_fill
        time_cell.border = border
        
        # Align whole hours (:00) to the right, half hours (:30) to the left
        if time_value and isinstance(time_value, str):
            if time_value.endswith(":00"):
                time_cell.alignment = Alignment(horizontal="right", vertical="center")
            elif time_value.endswith(":30"):
                time_cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                time_cell.alignment = center_alignment
        else:
            time_cell.alignment = center_alignment
        
        # Format lane columns with text wrapping
        for col_idx in range(2, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = lane_font
            cell.alignment = wrap_alignment  # Enable text wrapping
            cell.border = border
            
            lane_number = col_idx - 1  # Convert column index to lane number (1-8)
            
            # Add darker background for specific lane/time combinations (only for empty lanes)
            if time_value and isinstance(time_value, str):
                # Lanes 1-4 darker on half-hour times (:30)
                if time_value.endswith(":30") and lane_number in [1, 2, 3, 4]:
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                # Lanes 5-8 darker on whole-hour times (:00)
                elif time_value.endswith(":00") and lane_number in [5, 6, 7, 8]:
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")


wb_full = Workbook()
ws_full = wb_full.active
ws_full.title = "Bowling Planning"

# Write headers
headers = ["Tijd"] + [f"Baan {i}" for i in range(1, 9)]
ws_full.append(headers)

# Sort schedule by time for better readability
sorted_times = sorted(schedule.keys(), key=lambda x: datetime.strptime(x, "%H:%M"))

for time_str in sorted_times:
    row = [time_str] + [schedule[time_str][lane] for lane in range(1, 9)]
    ws_full.append(row)

# Apply formatting
format_worksheet(ws_full, schedule, "Full Schedule")

wb_full.save(output_file_full)
print(f"✅ Full schedule saved to {output_file_full}")

